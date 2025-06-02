# backend/blueprints/pdf_operations/pdf_to_excel_handler.py
import os
import uuid
import shutil
import pandas as pd
import tabula # tabula-py
from flask import current_app
from werkzeug.utils import secure_filename
from .utils import check_allowed_file, create_temp_folder, save_uploaded_file

ALLOWED_EXTENSIONS_PDF = {'pdf'}

def handle_pdf_to_excel(request_files, request_form):
    if 'files' not in request_files:
        return {'success': False, 'error': 'No file part in the request'}, 400
    
    file_stream = request_files.getlist('files')[0]
    if not file_stream or not file_stream.filename:
        raise ValueError('No file selected for PDF to Excel.')
    
    original_filename = file_stream.filename
    if not check_allowed_file(original_filename, ALLOWED_EXTENSIONS_PDF):
        raise ValueError(f"Invalid file type: {original_filename}. Only PDF files are allowed.")

    original_filename_secure = secure_filename(original_filename)
    request_temp_folder = create_temp_folder("pdf2excel_temp")
    
    try:
        temp_input_filepath = save_uploaded_file(file_stream, request_temp_folder)
        
        # Extract tables from PDF.
        # tabula.read_pdf returns a list of pandas DataFrames, one for each table found.
        # The 'pages' argument can be 'all', a page number, or a list of page numbers.
        # 'lattice=True' or 'stream=True' can be used depending on table structure.
        # This might need more advanced options or trying both lattice and stream if results are poor.
        try:
            dfs = tabula.read_pdf(temp_input_filepath, pages='all', multiple_tables=True, lattice=True)
            if not dfs: # If lattice didn't find tables, try stream mode
                 current_app.logger.info(f"Lattice mode found no tables in {original_filename_secure}, trying stream mode.")
                 dfs = tabula.read_pdf(temp_input_filepath, pages='all', multiple_tables=True, stream=True)

        except Exception as tabula_error:
            # tabula-py can raise various errors, including if Java is not found
            if "java.io.IOException" in str(tabula_error) or "JavaNotFoundError" in str(tabula_error):
                 raise FileNotFoundError("Java runtime not found or tabula setup issue. Please ensure Java is installed and accessible.") from tabula_error
            current_app.logger.error(f"Tabula PDF parsing error for {original_filename_secure}: {tabula_error}")
            raise Exception(f"Could not extract tables from PDF. The PDF might not contain detectable tables or is corrupted. Error: {str(tabula_error)}")


        if not dfs: # If still no DataFrames
            raise ValueError("No tables found in the PDF or tables are not in a detectable format.")

        output_filename_base = os.path.splitext(original_filename_secure)[0]
        output_excel_filename = f"{output_filename_base}_{uuid.uuid4().hex[:6]}.xlsx"
        output_excel_filepath = os.path.join(current_app.config['CONVERTED_FILES_FOLDER'], output_excel_filename)
        
        # Save all found tables into different sheets of the same Excel file
        with pd.ExcelWriter(output_excel_filepath, engine='openpyxl') as writer:
            for i, df in enumerate(dfs):
                if not df.empty:
                    df.to_excel(writer, sheet_name=f'Table_{i+1}', index=False)
                else:
                    current_app.logger.info(f"Skipping empty DataFrame (table {i+1}) from {original_filename_secure}")
        
        if not os.path.exists(output_excel_filepath) or os.path.getsize(output_excel_filepath) == 0:
             # Check if any sheets were actually written
            from openpyxl import load_workbook
            try:
                wb = load_workbook(output_excel_filepath)
                if not wb.sheetnames: # No sheets means no data was written
                    raise Exception("PDF to Excel conversion resulted in an empty Excel file. No valid tables might have been extracted.")
            except: # If file is truly empty or corrupt, load_workbook might fail
                raise Exception("PDF to Excel conversion failed to produce a valid Excel file.")


        response_data = {
            'success': True, 
            'message': f"Successfully extracted tables from '{original_filename_secure}' to Excel.",
            'download_url': f'/api/download/{output_excel_filename}', 
            'filename': output_excel_filename
        }
        return response_data, 200
    finally:
        if request_temp_folder and os.path.exists(request_temp_folder):
            try:
                shutil.rmtree(request_temp_folder)
            except OSError as e_clean:
                current_app.logger.error(f"Error cleaning temp folder {request_temp_folder} for pdf_to_excel: {e_clean}")